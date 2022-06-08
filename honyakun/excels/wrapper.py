from typing import Optional
import openpyxl


class ExcelWrapper():
    def __init__(self, filename: str, sheet_name: str) -> None:
        self.filename = filename
        self.wb = openpyxl.load_workbook(filename)
        self.ws = self.wb[sheet_name]

    def change_sheet(self, sheet_name: str) -> None:
        self.ws = self.wb[sheet_name]

    def read_column_texts(self, column: str, start: int, end: int) -> list[Optional[str]]:
        taples = self.ws[f"{column}{start}":f"{column}{end}"]
        result = [taple[0].value for taple in taples]
        return result

    def write_column_texts(self, column: str, start: int, texts: list[Optional[str]]) -> None:
        for i, text in enumerate(texts):
            excel_index = start + i
            self.ws[f"{column}{excel_index}"] = text

    def save(self, filename=None) -> None:
        if filename is None:
            return self.wb.save(filename=self.filename)
        self.wb.save(filename)
